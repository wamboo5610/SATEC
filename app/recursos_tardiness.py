"""Cálculo de tardanzas cruzando listado RRHH (lista) con reportes Excel del reloj."""

from __future__ import annotations

import os
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from . import attendance_logic as al
from . import database as db
from . import excel_import as xls

DEFAULT_LISTA_DIR = Path(r"C:\Users\ASUS\Downloads\recursos\lista")
DEFAULT_REPORTES_DIR = Path(r"C:\Users\ASUS\Downloads\recursos\reportes")

LISTA_HEADER_MARKERS = {"n°", "nº", "n", "apellidos y nombres", "apellidos y nombres ", "dni"}


def get_lista_dir() -> Path:
    custom = (os.environ.get("SATEC_RECURSOS_LISTA_DIR") or os.environ.get("SISAT_RECURSOS_LISTA_DIR") or "").strip()
    return Path(custom) if custom else DEFAULT_LISTA_DIR


def get_reportes_dir() -> Path:
    custom = (os.environ.get("SATEC_RECURSOS_REPORTES_DIR") or os.environ.get("SISAT_RECURSOS_REPORTES_DIR") or "").strip()
    return Path(custom) if custom else DEFAULT_REPORTES_DIR


def _project_root() -> Path:
    return Path(__file__).resolve().parent.parent


def get_folder_presets() -> dict:
    """Carpetas sugeridas para lista y reportes (solo las que existen en disco)."""
    root = _project_root()
    lista_candidates = [
        ("Downloads · recursos/lista", DEFAULT_LISTA_DIR),
        ("Proyecto · REPORTES", root / "REPORTES"),
        ("Proyecto · lista", root / "lista"),
        ("Downloads · recursos", DEFAULT_LISTA_DIR.parent),
    ]
    reportes_candidates = [
        ("Downloads · recursos/reportes", DEFAULT_REPORTES_DIR),
        ("Proyecto · REPORTES", root / "REPORTES"),
        ("Proyecto · reportes", root / "reportes"),
        ("Downloads · recursos", DEFAULT_REPORTES_DIR.parent),
    ]

    def _build_options(candidates: list[tuple[str, Path]]) -> list[dict]:
        seen: set[str] = set()
        options: list[dict] = []
        for label, path in candidates:
            resolved = path.resolve()
            key = str(resolved).lower()
            if key in seen:
                continue
            seen.add(key)
            options.append({
                "label": label,
                "path": str(resolved),
                "exists": resolved.exists(),
                "xlsx_count": len(list(resolved.glob("*.xlsx"))) if resolved.exists() else 0,
            })
        return options

    return {
        "lista": _build_options(lista_candidates),
        "reportes": _build_options(reportes_candidates),
        "default_lista_dir": str(get_lista_dir()),
        "default_reportes_dir": str(get_reportes_dir()),
    }


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _normalize_dni(value) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits.lstrip("0") or "0"


def _normalize_name(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").upper().strip())


def _name_match_score(left: str, right: str) -> float:
    tokens_left = set(_normalize_name(left).split())
    tokens_right = set(_normalize_name(right).split())
    if not tokens_left or not tokens_right:
        return 0.0
    return len(tokens_left & tokens_right) / max(len(tokens_left), len(tokens_right))


def _find_lista_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | tuple[None, None]:
    for idx, row in enumerate(rows):
        mapping: dict[str, int] = {}
        for col, cell in enumerate(row):
            key = _normalize_header(cell)
            if key in LISTA_HEADER_MARKERS or "apellidos" in key or key == "dni":
                mapping[key] = col
        if any("apellidos" in key for key in mapping) and "dni" in mapping:
            return idx, mapping
    return None, None


def _col(mapping: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in mapping:
            return mapping[name]
    for key, col in mapping.items():
        for name in names:
            if name in key:
                return col
    return None


def parse_person_list_bytes(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return parse_person_list_rows(rows)


def parse_person_list_file(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return parse_person_list_rows(rows)


def parse_person_list_rows(rows: list[tuple]) -> list[dict]:
    header_idx, mapping = _find_lista_header_row(rows)
    if header_idx is None or mapping is None:
        raise ValueError("No se encontró encabezado en la lista (APELLIDOS y NOMBRES, DNI)")

    col_num = _col(mapping, "n°", "nº", "n")
    col_name = _col(mapping, "apellidos y nombres", "apellidos y nombres ")
    col_dni = _col(mapping, "dni")
    if col_name is None or col_dni is None:
        raise ValueError("La lista debe incluir columnas APELLIDOS y NOMBRES y DNI")

    people: list[dict] = []
    for row in rows[header_idx + 1:]:
        if not row or col_name >= len(row) or col_dni >= len(row):
            continue
        name = str(row[col_name] or "").strip()
        dni_raw = str(row[col_dni] or "").strip()
        if not name or not dni_raw:
            continue
        num = row[col_num] if col_num is not None and col_num < len(row) else len(people) + 1
        people.append({
            "num": num,
            "name": name,
            "name_norm": _normalize_name(name),
            "dni": _normalize_dni(dni_raw),
            "dni_raw": dni_raw,
        })
    if not people:
        raise ValueError("No se encontraron personas en la lista")
    return people


def list_lista_files(lista_dir: Path | None = None) -> list[dict]:
    folder = lista_dir or get_lista_dir()
    if not folder.exists():
        return []
    files = []
    for path in sorted(folder.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
        stat = path.stat()
        files.append({
            "filename": path.name,
            "size_kb": round(stat.st_size / 1024, 1),
            "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
        })
    return files


def find_lista_file(lista_dir: Path | None = None, filename: str | None = None) -> Path | None:
    folder = lista_dir or get_lista_dir()
    if not folder.exists():
        return None
    if filename:
        path = folder / Path(filename).name
        return path if path.exists() and path.suffix.lower() == ".xlsx" else None
    files = sorted(folder.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def list_report_files(reportes_dir: Path | None = None) -> list[dict]:
    folder = reportes_dir or get_reportes_dir()
    if not folder.exists():
        return []
    files = []
    for path in sorted(folder.glob("*.xlsx")):
        stat = path.stat()
        files.append({
            "filename": path.name,
            "size_kb": round(stat.st_size / 1024, 1),
            "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            "sede_guess": infer_sede_from_filename(path.name),
        })
    return files


REPORT_FILENAME_SEDE = (
    ("RELLENO", "Relleno Sanitario"),
    ("PALACIO", "Sede Principal"),
)

FILE_SOURCE_PRIORITY = {
    "relleno": 30,
    "palacio": 20,
    "sedes": 10,
    "otros": 0,
}


def _file_source_key(filename: str) -> str:
    stem = Path(filename).stem.upper()
    if "RELLENO" in stem:
        return "relleno"
    if "PALACIO" in stem:
        return "palacio"
    if "SEDES" in stem:
        return "sedes"
    return "otros"


def _match_filename_sede(filename: str) -> str | None:
    stem = Path(filename).stem.upper()
    for token, sede_name in REPORT_FILENAME_SEDE:
        if token in stem:
            return sede_name
    return None


def build_user_sede_hints() -> dict[str, int]:
    """Sede más frecuente por empleado según relojes / importaciones previas en SATEC."""
    counts: dict[str, dict[int, int]] = {}
    device_map = db.get_device_serial_map()
    with db.get_conn() as conn:
        rows = conn.execute("SELECT user_id, device_serial FROM users_cache").fetchall()
    for row in rows:
        uid = str(row["user_id"])
        serial = str(row["device_serial"] or "")
        sid = None
        match = re.match(r"EXCEL-S(\d+)", serial, re.I)
        if match:
            sid = int(match.group(1))
        elif serial in device_map and device_map[serial].get("sede_id"):
            sid = device_map[serial]["sede_id"]
        if sid:
            counts.setdefault(uid, {})
            counts[uid][sid] = counts[uid].get(sid, 0) + 1
    return {uid: max(sede_counts, key=sede_counts.get) for uid, sede_counts in counts.items()}


def resolve_report_sede(
    filename: str,
    sede_name_to_id: dict[str, int],
    user_sede_hints: dict[str, int],
    user_id: str | None = None,
) -> tuple[int | None, str]:
    """Resuelve sede SATEC para un reporte (por archivo o por empleado en sedes varias)."""
    sede_names = {name.lower(): (sid, name) for name, sid in sede_name_to_id.items()}
    mapped = _match_filename_sede(filename)
    if mapped:
        key = mapped.lower()
        if key in sede_names:
            sid, name = sede_names[key]
            return sid, name
        for db_name, (sid, canonical) in sede_names.items():
            if key in db_name or db_name in key:
                return sid, canonical

    if user_id and _file_source_key(filename) == "sedes":
        hint_sid = user_sede_hints.get(str(user_id))
        if hint_sid:
            for name, sid in sede_name_to_id.items():
                if sid == hint_sid:
                    return sid, name

    if user_id and user_id in user_sede_hints:
        hint_sid = user_sede_hints[user_id]
        for name, sid in sede_name_to_id.items():
            if sid == hint_sid:
                return sid, name

    if _file_source_key(filename) == "sedes":
        if user_id:
            for preferred in ("Sede Principal", "Maestranza", "Terminal", "Galerias", "Estadio"):
                if preferred in sede_name_to_id:
                    return sede_name_to_id[preferred], preferred
        return None, "Sedes varias (por empleado)"

    for preferred in ("Sede Principal", "Maestranza"):
        if preferred in sede_name_to_id:
            return sede_name_to_id[preferred], preferred
    if sede_name_to_id:
        fallback_name = sorted(sede_name_to_id)[0]
        return sede_name_to_id[fallback_name], fallback_name
    return None, infer_sede_label(filename)


def infer_sede_label(filename: str) -> str:
    mapped = _match_filename_sede(filename)
    if mapped:
        return mapped
    stem = Path(filename).stem.upper()
    if "SEDES" in stem:
        return "Sedes MPO"
    cleaned = re.sub(r"REPORTE|MES|DE|DEL|LA|EL|\d{4}", "", stem, flags=re.I)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -_")
    return cleaned.title() if cleaned else "Reporte"


def infer_sede_from_filename(filename: str) -> str:
    return infer_sede_label(filename)


def _attach_schedule_summary(meta: dict, schedules_by_sede: dict[int, dict]) -> None:
    from . import calendar_rules as cr

    rows = []
    for sid, sched in sorted(schedules_by_sede.items(), key=lambda x: x[1].get("sede_name", "")):
        rows.append({
            "sede_id": sid,
            "sede_name": sched.get("sede_name", f"Sede {sid}"),
            "entry_time": sched.get("entry_time"),
            "exit_time": sched.get("exit_time"),
            "lunch_start": sched.get("lunch_start"),
            "lunch_end": sched.get("lunch_end"),
            "grace_minutes": sched.get("grace_minutes", 0),
            "lunch_grace_minutes": sched.get("lunch_grace_minutes", 0),
            "work_days": cr.work_days_label(sched),
        })
    meta["schedules_applied"] = rows
    meta["monthly_tolerance_minutes"] = al.MONTHLY_TOLERANCE_MINUTES
    meta["uses_satec_rules"] = True


def tag_records_with_sede(
    records: list[dict],
    *,
    filename: str,
    sede_name_to_id: dict[str, int],
    user_sede_hints: dict[str, int],
    force_sede_id: int | None = None,
    force_sede_name: str | None = None,
) -> list[dict]:
    source_key = _file_source_key(filename)
    priority = FILE_SOURCE_PRIORITY.get(source_key, 0)
    tagged = []
    for rec in records:
        row = dict(rec)
        uid = str(row.get("user_id", "")).strip()
        if force_sede_id:
            sid = force_sede_id
            sname = force_sede_name or next(
                (n for n, i in sede_name_to_id.items() if i == force_sede_id), f"Sede {force_sede_id}"
            )
        else:
            sid, sname = resolve_report_sede(filename, sede_name_to_id, user_sede_hints, uid)
        row["sede_id"] = sid
        row["sede_name"] = sname
        row["source_file"] = filename
        row["source_priority"] = priority
        row["device_serial"] = f"RECURSOS-S{sid or 0}-{source_key.upper()}"
        row["source"] = "recursos"
        tagged.append(row)
    return tagged


def resolve_duplicate_days(records: list[dict]) -> list[dict]:
    """Si un empleado tiene el mismo día en varios reportes, conserva el de sede más específica."""
    by_day: dict[tuple[str, str], list[dict]] = {}
    for rec in records:
        day = al.extract_date(rec.get("timestamp", ""))
        uid = str(rec.get("user_id", "")).strip()
        if uid and day:
            by_day.setdefault((uid, day), []).append(rec)

    kept: list[dict] = []
    for day_rows in by_day.values():
        source_files = {r.get("source_file") for r in day_rows}
        if len(source_files) <= 1:
            kept.extend(day_rows)
            continue
        best_priority = max(int(r.get("source_priority", 0) or 0) for r in day_rows)
        kept.extend(r for r in day_rows if int(r.get("source_priority", 0) or 0) == best_priority)
    return sorted(kept, key=lambda x: (x.get("timestamp", ""), x.get("user_id", "")))


def build_report_user_index(records: list[dict]) -> dict[str, dict]:
    users: dict[str, dict] = {}
    for rec in records:
        uid = str(rec.get("user_id", "")).strip()
        if not uid:
            continue
        if uid not in users:
            users[uid] = {
                "user_id": uid,
                "user_name": rec.get("user_name") or "",
                "name_norm": _normalize_name(rec.get("user_name") or ""),
                "dni_norm": _normalize_dni(uid),
            }
    return users


def match_person_to_user(person: dict, users: dict[str, dict]) -> tuple[dict | None, str]:
    for user in users.values():
        if person["dni"] == user["dni_norm"] and len(person["dni"]) >= 6:
            return user, "dni"
    best_user = None
    best_score = 0.0
    for user in users.values():
        score = _name_match_score(person["name_norm"], user["name_norm"])
        if score > best_score:
            best_score = score
            best_user = user
    if best_user and best_score >= 0.5:
        return best_user, f"name:{best_score:.2f}"
    return None, "none"


def build_person_matches(people: list[dict], users: dict[str, dict]) -> list[dict]:
    matches = []
    for person in people:
        user, method = match_person_to_user(person, users)
        matches.append({
            **person,
            "matched": user is not None,
            "match_method": method,
            "report_user_id": user["user_id"] if user else None,
            "report_user_name": user["user_name"] if user else None,
        })
    return matches


def read_report_bytes(reportes_dir: Path, filename: str) -> bytes:
    safe_name = Path(filename).name
    path = reportes_dir / safe_name
    if not path.exists() or path.suffix.lower() != ".xlsx":
        raise FileNotFoundError(f"Reporte no encontrado: {safe_name}")
    return path.read_bytes()


def _empty_person_row(match: dict) -> dict:
    uid = match.get("report_user_id") or match.get("dni_raw") or ""
    return {
        "user_id": uid,
        "name": match.get("name") or match.get("report_user_name") or "",
        "sede": "—",
        "dni": match.get("dni_raw"),
        "lista_num": match.get("num"),
        "days_with_attendance": 0,
        "punctual_days": 0,
        "late_days": 0,
        "late_days_entrada": 0,
        "late_days_almuerzo": 0,
        "late_minutes_gross": 0,
        "late_minutes": 0,
        "late_minutes_entrada": 0,
        "late_minutes_almuerzo": 0,
        "monthly_tolerance_applied": 0,
        "avg_late_minutes": 0,
        "non_work_days": 0,
        "sin_marcaciones": True,
    }


def restrict_report_to_lista(
    aggregated: dict,
    matches: list[dict],
    *,
    lista_persons: int,
) -> dict:
    """Deja en el reporte únicamente las personas del listado Excel (incluye las sin marcaciones)."""
    lista_matches = [m for m in matches if m.get("matched")]
    allowed_ids = {str(m["report_user_id"]) for m in lista_matches if m.get("report_user_id")}

    computed_by_id = {str(p.get("user_id")): p for p in aggregated.get("by_person", [])}
    details = [
        d for d in aggregated.get("details", [])
        if str(d.get("user_id")) in allowed_ids
    ]

    by_person: list[dict] = []
    for match in sorted(lista_matches, key=lambda m: (m.get("num") or 0, m.get("name") or "")):
        uid = str(match.get("report_user_id") or "")
        row = dict(computed_by_id.get(uid) or _empty_person_row(match))
        row["user_id"] = uid or row.get("user_id")
        row["name"] = match.get("name") or row.get("name") or ""
        row["dni"] = match.get("dni_raw")
        row["lista_num"] = match.get("num")
        if uid and uid not in computed_by_id:
            row["sin_marcaciones"] = True
        else:
            row.pop("sin_marcaciones", None)
        by_person.append(row)

    laborable = [d for d in details if d.get("dia_laborable", True)]
    total_late_minutes_gross = sum(p.get("late_minutes_gross", 0) or 0 for p in by_person)
    total_tolerance_applied = sum(p.get("monthly_tolerance_applied", 0) or 0 for p in by_person)
    total_late_minutes_net = sum(p.get("late_minutes", 0) or 0 for p in by_person)

    filtered = dict(aggregated)
    filtered["details"] = details
    filtered["by_person"] = by_person
    filtered["total_days"] = len(laborable)
    filtered["late_days"] = sum(1 for d in laborable if (d.get("tardanza_total_minutos") or 0) > 0)
    filtered["total_late_minutes_gross"] = total_late_minutes_gross
    filtered["total_tolerance_applied"] = total_tolerance_applied
    filtered["total_late_minutes"] = total_late_minutes_net
    filtered["total_persons"] = len(by_person)
    filtered["persons_with_late"] = sum(1 for p in by_person if (p.get("late_minutes") or 0) > 0)
    filtered["non_work_days"] = len(details) - len(laborable)
    filtered["lista_only"] = True
    filtered["meta"] = {
        **(aggregated.get("meta") or {}),
        "lista_only": True,
        "lista_persons": lista_persons,
        "shown_persons": len(by_person),
    }
    return filtered


def calculate_recursos_tardiness(
    *,
    lista_path: Path | None = None,
    lista_dir: Path | None = None,
    lista_filename: str | None = None,
    lista_bytes: bytes | None = None,
    reportes_dir: Path | None = None,
    report_files: list[str] | None = None,
    sede_id: int | None = None,
    schedule: dict | None = None,
    default_sede_name: str | None = None,
) -> dict:
    lista_folder = lista_dir or (lista_path.parent if lista_path else None) or get_lista_dir()
    reportes_folder = reportes_dir or get_reportes_dir()
    if not reportes_folder.exists():
        raise FileNotFoundError(f"Carpeta de reportes no encontrada: {reportes_folder}")

    sede_name_to_id = db.get_sede_name_to_id()
    schedules_by_sede_id = db.get_schedules_by_sede_id()
    user_sede_hints = build_user_sede_hints()
    force_sede_name = default_sede_name
    if sede_id and not force_sede_name:
        force_sede_name = next((n for n, sid in sede_name_to_id.items() if sid == sede_id), None)

    if lista_bytes:
        people = parse_person_list_bytes(lista_bytes)
        lista_file = "lista_subida.xlsx"
    else:
        lista_file_path = lista_path or find_lista_file(lista_folder, lista_filename)
        if not lista_file_path or not lista_file_path.exists():
            raise FileNotFoundError(f"No se encontró listado en {lista_folder}")
        people = parse_person_list_file(lista_file_path)
        lista_file = lista_file_path.name

    available_reports = list_report_files(reportes_folder)
    selected = report_files or [item["filename"] for item in available_reports]
    if not selected:
        raise ValueError("No hay reportes Excel en la carpeta configurada")

    all_records: list[dict] = []
    files_meta: list[dict] = []
    users_index: dict[str, dict] = {}

    for filename in selected:
        content = read_report_bytes(reportes_folder, filename)
        sede_label = infer_sede_from_filename(filename)
        records, meta = xls.parse_zk_excel(content, sede_name=sede_label, source_label="recursos")
        tagged = tag_records_with_sede(
            records,
            filename=filename,
            sede_name_to_id=sede_name_to_id,
            user_sede_hints=user_sede_hints,
            force_sede_id=sede_id if schedule else None,
            force_sede_name=force_sede_name,
        )
        users_index.update(build_report_user_index(tagged))
        all_records.extend(tagged)
        sample_sid = tagged[0]["sede_id"] if tagged else None
        sample_sname = tagged[0]["sede_name"] if tagged else sede_label
        files_meta.append({
            "filename": filename,
            "sede_name": sample_sname,
            "sede_id": sample_sid,
            "records": meta.get("total_records", len(records)),
            "persons": meta.get("total_persons", 0),
            "date_from": meta.get("date_from"),
            "date_to": meta.get("date_to"),
        })

    matches = build_person_matches(people, users_index)
    matched_ids = {m["report_user_id"] for m in matches if m["report_user_id"]}
    matched_by_id = {m["report_user_id"]: m for m in matches if m["report_user_id"]}

    filtered_records = []
    for rec in all_records:
        uid = str(rec.get("user_id", "")).strip()
        if uid not in matched_ids:
            continue
        person = matched_by_id[uid]
        rec = dict(rec)
        rec["user_name"] = rec.get("user_name") or person["name"]
        rec["lista_dni"] = person["dni_raw"]
        rec["lista_num"] = person["num"]
        filtered_records.append(rec)

    if not filtered_records:
        unmatched = [m for m in matches if not m["matched"]]
        raise ValueError(
            "Ninguna persona de la lista tiene marcaciones en los reportes seleccionados. "
            f"Sin coincidencia: {len(unmatched)} de {len(matches)}."
        )

    resolved = resolve_duplicate_days(filtered_records)
    deduped = al.dedupe_attendance(resolved)

    if schedule and sede_id:
        forced = al.normalize_schedule(schedule)
        schedules_by_sede_id = dict(schedules_by_sede_id)
        schedules_by_sede_id[sede_id] = {**schedules_by_sede_id.get(sede_id, {}), **forced, "sede_id": sede_id}
        for rec in deduped:
            rec["sede_id"] = sede_id
            if force_sede_name:
                rec["sede_name"] = force_sede_name

    emp_schedules = db.get_schedules_by_user_id()
    global_dates, by_sede = db.get_holiday_dates_lookup()
    report = al.build_tardiness_report(
        al.infer_punch_types(deduped),
        schedules_by_sede_id,
        sede_name_to_id,
        emp_schedules,
        global_dates,
        by_sede,
    )
    aggregated = restrict_report_to_lista(
        al.aggregate_tardiness_report(report),
        matches,
        lista_persons=len(people),
    )

    dates = sorted({rec["timestamp"][:10] for rec in deduped if rec.get("timestamp")})
    persons_with_data = {rec["user_id"] for rec in deduped}
    matched_without_data = [
        {
            "num": m["num"],
            "name": m["name"],
            "dni": m["dni_raw"],
            "report_user_id": m["report_user_id"],
            "report_user_name": m["report_user_name"],
        }
        for m in matches
        if m["matched"] and m["report_user_id"] not in persons_with_data
    ]
    unmatched_people = [
        {"num": m["num"], "name": m["name"], "dni": m["dni_raw"]}
        for m in matches
        if not m["matched"]
    ]

    aggregated["meta"].update({
        "lista_file": lista_file,
        "lista_persons": len(people),
        "matched_persons": sum(1 for m in matches if m["matched"]),
        "persons_with_attendance": len(persons_with_data),
        "unmatched_persons": len(unmatched_people),
        "matched_without_attendance": len(matched_without_data),
        "report_files": files_meta,
        "records_total": len(all_records),
        "records_filtered": len(filtered_records),
        "records_deduped": len(deduped),
        "date_from": dates[0] if dates else None,
        "date_to": dates[-1] if dates else None,
        "sede_name": force_sede_name or "Automático por sede",
        "schedule_mode": "forced" if schedule and sede_id else "automatic",
        "lista_dir": str(lista_folder),
        "reportes_dir": str(reportes_folder),
    })
    _attach_schedule_summary(aggregated["meta"], schedules_by_sede_id)
    aggregated["matching"] = {
        "matched": [
            {
                "num": m["num"],
                "name": m["name"],
                "dni": m["dni_raw"],
                "report_user_id": m["report_user_id"],
                "report_user_name": m["report_user_name"],
                "match_method": m["match_method"],
            }
            for m in matches if m["matched"]
        ],
        "unmatched": unmatched_people,
        "without_attendance": matched_without_data,
    }
    return aggregated


def get_recursos_status(
    lista_dir: Path | None = None,
    reportes_dir: Path | None = None,
) -> dict:
    from . import calendar_rules as cr

    lista_folder = lista_dir or get_lista_dir()
    reportes_folder = reportes_dir or get_reportes_dir()
    lista_file = find_lista_file(lista_folder)
    schedules_by_sede_id = db.get_schedules_by_sede_id()
    schedules = []
    for sid, sched in sorted(schedules_by_sede_id.items(), key=lambda x: x[1].get("sede_name", "")):
        schedules.append({
            "sede_id": sid,
            "sede_name": sched.get("sede_name", f"Sede {sid}"),
            "entry_time": sched.get("entry_time"),
            "exit_time": sched.get("exit_time"),
            "lunch_start": sched.get("lunch_start"),
            "lunch_end": sched.get("lunch_end"),
            "grace_minutes": sched.get("grace_minutes", 0),
            "work_days": cr.work_days_label(sched),
        })
    global_dates, _by_sede = db.get_holiday_dates_lookup()
    report_files = list_report_files(reportes_folder)
    lista_files = list_lista_files(lista_folder)
    sede_name_to_id = db.get_sede_name_to_id()
    for item in report_files:
        sid, sname = resolve_report_sede(item["filename"], sede_name_to_id, {})
        item["sede_id"] = sid
        item["sede_name_resolved"] = sname
    return {
        "lista_dir": str(lista_folder),
        "reportes_dir": str(reportes_folder),
        "lista_dir_exists": lista_folder.exists(),
        "reportes_dir_exists": reportes_folder.exists(),
        "lista_file": lista_file.name if lista_file else None,
        "lista_file_exists": bool(lista_file and lista_file.exists()),
        "lista_files": lista_files,
        "report_files": report_files,
        "folder_presets": get_folder_presets(),
        "schedules": schedules,
        "holiday_count": len(global_dates),
        "monthly_tolerance_minutes": al.MONTHLY_TOLERANCE_MINUTES,
        "report_sede_mapping": [
            {"pattern": token, "sede_name": sede_name}
            for token, sede_name in REPORT_FILENAME_SEDE
        ],
    }