import re
from collections import defaultdict
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import attendance_logic as al

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=12, color="1E3A5F")
HIGHLIGHT_GROSS_FILL = PatternFill("solid", fgColor="FFF3CD")
HIGHLIGHT_TOL_FILL = PatternFill("solid", fgColor="CCE5FF")
HIGHLIGHT_NET_FILL = PatternFill("solid", fgColor="D4EDDA")
HIGHLIGHT_HEADER_GROSS = PatternFill("solid", fgColor="E6A800")
HIGHLIGHT_HEADER_TOL = PatternFill("solid", fgColor="2E75B6")
HIGHLIGHT_HEADER_NET = PatternFill("solid", fgColor="548235")
TOTAL_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")
TOTAL_FONT = Font(bold=True, size=11, color="1E3A5F")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TARDINESS_MIN_COLS = {
    "gross": 11,
    "tolerance": 12,
    "net": 13,
}


def safe_sheet_name(user_id, name, used):
    base = re.sub(r'[\\/*?:\[\]]', '', f"{user_id} {name or 'Sin nombre'}")[:28]
    name_candidate = base
    n = 1
    while name_candidate in used:
        suffix = f"_{n}"
        name_candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name_candidate)
    return name_candidate


def enrich_row(r, device_map):
    serial = r.get("device_serial") or ""
    dev = device_map.get(serial, {})
    row = al.enrich_record(dict(r))
    row["sede_name"] = dev.get("sede_name", "Sin sede")
    row["device_name"] = dev.get("device_name", serial or "—")
    return row


def enrich_rows(rows, device_map):
    enriched = [enrich_row(r, device_map) for r in rows]
    return al.infer_punch_types(enriched)


def group_by_person(rows):
    groups = defaultdict(list)
    names = {}
    for r in rows:
        uid = str(r.get("user_id", ""))
        groups[uid].append(r)
        if r.get("user_name"):
            names[uid] = r["user_name"]
    for uid in groups:
        groups[uid].sort(key=lambda x: x.get("timestamp", ""))
    return groups, names


def build_summary_rows(groups, names):
    summary = []
    for uid, recs in sorted(groups.items(), key=lambda x: names.get(x[0], x[0])):
        summary.append({
            "user_id": uid,
            "user_name": names.get(uid, ""),
            "sede_name": recs[0].get("sede_name", "—") if recs else "—",
            "device_name": recs[0].get("device_name", "—") if recs else "—",
            "total": len(recs),
            "first": recs[0].get("timestamp", "") if recs else "",
            "last": recs[-1].get("timestamp", "") if recs else "",
        })
    return summary


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def tardiness_report_totals(report_data: dict) -> tuple[int, int, int, int]:
    gross = int(report_data.get("total_late_minutes_gross", 0) or 0)
    applied = report_data.get("total_tolerance_applied")
    if applied is None:
        applied = sum(int(p.get("monthly_tolerance_applied", 0) or 0) for p in report_data.get("by_person", []))
    applied = int(applied or 0)
    net = int(report_data.get("total_late_minutes", 0) or 0)
    policy = int(report_data.get("monthly_tolerance_minutes", 20) or 20)
    return gross, applied, net, policy


def _style_highlight_value(cell, fill, bold=True, size=11):
    cell.fill = fill
    cell.font = Font(bold=bold, size=size, color="1E3A5F")
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_tardiness_highlight_totals(ws, start_row: int, report_data: dict) -> int:
    gross, applied, net, policy = tardiness_report_totals(report_data)
    ws.cell(row=start_row, column=1, value="TOTALES DE TARDANZA DEL PERÍODO").font = Font(
        bold=True, size=12, color="1E3A5F"
    )
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)

    blocks = [
        (start_row + 1, 1, "Total tardanza (bruto)", gross, HIGHLIGHT_GROSS_FILL),
        (start_row + 1, 4, "Total tolerancia aplicada", applied, HIGHLIGHT_TOL_FILL),
        (start_row + 1, 7, "Total tardanza (neto)", net, HIGHLIGHT_NET_FILL),
    ]
    for row, col, label, value, fill in blocks:
        lbl = ws.cell(row=row, column=col, value=label)
        lbl.font = Font(bold=True, size=10, color="1E3A5F")
        lbl.alignment = Alignment(horizontal="right", vertical="center")
        val = ws.cell(row=row, column=col + 1, value=value)
        _style_highlight_value(val, fill, bold=True, size=12)
        unit = ws.cell(row=row, column=col + 2, value="min")
        unit.font = Font(size=9, color="666666")
        unit.alignment = Alignment(horizontal="left", vertical="center")

    note = ws.cell(
        row=start_row + 2,
        column=1,
        value=f"Política RRHH: hasta {policy} min de tolerancia mensual por empleado (neto = bruto − tolerancia aplicada)",
    )
    note.font = Font(italic=True, size=9, color="666666")
    ws.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 2, end_column=10)
    return start_row + 3


def _style_tardiness_minute_headers(ws, row: int):
    styles = {
        TARDINESS_MIN_COLS["gross"]: (HIGHLIGHT_HEADER_GROSS, Font(bold=True, color="FFFFFF", size=11)),
        TARDINESS_MIN_COLS["tolerance"]: (HIGHLIGHT_HEADER_TOL, Font(bold=True, color="FFFFFF", size=11)),
        TARDINESS_MIN_COLS["net"]: (HIGHLIGHT_HEADER_NET, Font(bold=True, color="FFFFFF", size=11)),
    }
    for col, (fill, font) in styles.items():
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _style_tardiness_minute_row(ws, row: int, *, total_row: bool = False):
    fills = {
        TARDINESS_MIN_COLS["gross"]: HIGHLIGHT_GROSS_FILL if total_row else PatternFill("solid", fgColor="FFFBEB"),
        TARDINESS_MIN_COLS["tolerance"]: HIGHLIGHT_TOL_FILL if total_row else PatternFill("solid", fgColor="EBF5FF"),
        TARDINESS_MIN_COLS["net"]: HIGHLIGHT_NET_FILL if total_row else PatternFill("solid", fgColor="E8F5E9"),
    }
    for col, fill in fills.items():
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = TOTAL_FONT if total_row else Font(bold=True, size=10, color="1E3A5F")
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _append_tardiness_person_totals_row(ws, row: int, report_data: dict):
    gross, applied, net, _policy = tardiness_report_totals(report_data)
    ws.cell(row=row, column=1, value="TOTAL GENERAL").font = TOTAL_FONT
    for col in range(2, TARDINESS_MIN_COLS["gross"]):
        ws.cell(row=row, column=col, value="").border = BORDER
    ws.cell(row=row, column=TARDINESS_MIN_COLS["gross"], value=gross)
    ws.cell(row=row, column=TARDINESS_MIN_COLS["tolerance"], value=applied)
    ws.cell(row=row, column=TARDINESS_MIN_COLS["net"], value=net)
    for col in range(TARDINESS_MIN_COLS["net"] + 1, 15):
        ws.cell(row=row, column=col, value="").border = BORDER
    _style_tardiness_minute_row(ws, row, total_row=True)
    for col in range(1, 15):
        if col not in TARDINESS_MIN_COLS.values():
            ws.cell(row=row, column=col).fill = TOTAL_ROW_FILL
            ws.cell(row=row, column=col).border = BORDER


def auto_width(ws, min_w=12, max_w=28):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, min_w), max_w)


def build_excel_by_person(rows, meta=None):
    meta = meta or {}
    wb = Workbook()
    used_names = set()

    # --- Hoja RESUMEN ---
    ws_sum = wb.active
    ws_sum.title = "RESUMEN"
    ws_sum["A1"] = "SISTEMA DE CONTROL DE ASISTENCIA — REPORTE POR PERSONA"
    ws_sum["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws_sum.merge_cells("A1:G1")
    info = [
        ("Generado:", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Desde:", meta.get("date_from", "Todos")),
        ("Hasta:", meta.get("date_to", "Todos")),
        ("Sede:", meta.get("sede_name", "Todas")),
        ("Reloj:", meta.get("device_name", "Todos")),
    ]
    for i, (k, v) in enumerate(info, 3):
        ws_sum.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws_sum.cell(row=i, column=2, value=v)

    groups, names = group_by_person(rows)
    summary = build_summary_rows(groups, names)

    hdr_row = 9
    headers = ["ID", "Nombre", "Sede", "Reloj", "Total Marcaciones", "Primera", "Última"]
    for c, h in enumerate(headers, 1):
        ws_sum.cell(row=hdr_row, column=c, value=h)
    style_header_row(ws_sum, hdr_row, len(headers))

    for i, s in enumerate(summary, hdr_row + 1):
        ws_sum.cell(row=i, column=1, value=s["user_id"])
        ws_sum.cell(row=i, column=2, value=s["user_name"])
        ws_sum.cell(row=i, column=3, value=s["sede_name"])
        ws_sum.cell(row=i, column=4, value=s["device_name"])
        ws_sum.cell(row=i, column=5, value=s["total"])
        ws_sum.cell(row=i, column=6, value=s["first"])
        ws_sum.cell(row=i, column=7, value=s["last"])

    auto_width(ws_sum)

    # --- Una hoja por persona ---
    detail_headers = ["Fecha", "Hora", "Tipo", "Verificación", "Sede", "Reloj", "Origen"]
    for uid, recs in sorted(groups.items(), key=lambda x: names.get(x[0], x[0])):
        sheet = safe_sheet_name(uid, names.get(uid), used_names)
        ws = wb.create_sheet(sheet)
        ws["A1"] = f"EMPLEADO: {names.get(uid, 'Sin nombre')}"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:G1")
        ws["A2"] = f"ID: {uid}"
        ws["A3"] = f"Total registros: {len(recs)}"

        row = 5
        for c, h in enumerate(detail_headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(detail_headers))

        for rec in recs:
            row += 1
            ts = rec.get("timestamp", "")
            parts = ts.split(" ") if ts else ["", ""]
            ws.cell(row=row, column=1, value=parts[0] if parts else "")
            ws.cell(row=row, column=2, value=parts[1] if len(parts) > 1 else "")
            ws.cell(row=row, column=3, value=rec.get("punch_type", ""))
            ws.cell(row=row, column=4, value=rec.get("verify_label", ""))
            ws.cell(row=row, column=5, value=rec.get("sede_name", ""))
            ws.cell(row=row, column=6, value=rec.get("device_name", ""))
            ws.cell(row=row, column=7, value=rec.get("source", ""))

        auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_excel_by_sede(rows, meta=None):
    meta = meta or {}
    wb = Workbook()
    used = set()
    by_sede = defaultdict(list)
    for r in rows:
        by_sede[r.get("sede_name", "Sin sede")].append(r)

    ws0 = wb.active
    ws0.title = "RESUMEN SEDES"
    ws0["A1"] = "SISTEMA DE CONTROL DE ASISTENCIA — REPORTE POR SEDE"
    ws0["A1"].font = Font(bold=True, size=14)
    headers = ["Sede", "Empleados", "Marcaciones"]
    for c, h in enumerate(headers, 1):
        ws0.cell(row=3, column=c, value=h)
    style_header_row(ws0, 3, 3)

    row = 4
    for sede in sorted(by_sede.keys()):
        recs = by_sede[sede]
        persons = len({r["user_id"] for r in recs})
        ws0.cell(row=row, column=1, value=sede)
        ws0.cell(row=row, column=2, value=persons)
        ws0.cell(row=row, column=3, value=len(recs))
        row += 1

    for sede, recs in sorted(by_sede.items()):
        sname = re.sub(r'[\\/*?:\[\]]', '', sede)[:31]
        while sname in used:
            sname = sname[:28] + "..."
        used.add(sname)
        ws = wb.create_sheet(sname)
        headers = ["ID", "Nombre", "Fecha/Hora", "Tipo", "Verificación", "Reloj"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)
        style_header_row(ws, 1, len(headers))
        for i, r in enumerate(sorted(recs, key=lambda x: (x.get("user_id", ""), x.get("timestamp", ""))), 2):
            ws.cell(row=i, column=1, value=r.get("user_id"))
            ws.cell(row=i, column=2, value=r.get("user_name"))
            ws.cell(row=i, column=3, value=r.get("timestamp"))
            ws.cell(row=i, column=4, value=r.get("punch_type"))
            ws.cell(row=i, column=5, value=r.get("verify_label"))
            ws.cell(row=i, column=6, value=r.get("device_name"))
        auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_meta_block(ws, title, meta, start_row=1):
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=14, color="1E3A5F")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
    info = [
        ("Generado:", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Desde:", meta.get("date_from", "Todos")),
        ("Hasta:", meta.get("date_to", "Todos")),
        ("Sede:", meta.get("sede_name", "Todas")),
        ("Reloj:", meta.get("device_name", "Todos")),
    ]
    for i, (k, v) in enumerate(info, start_row + 2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    return start_row + 2 + len(info)


def build_excel_tardiness(report_data, meta=None):
    meta = meta or {}
    wb = Workbook()
    used_names = set()

    ws_sum = wb.active
    ws_sum.title = "RESUMEN EMPLEADOS"
    meta_end = _write_meta_block(ws_sum, "REPORTE DE TARDANZAS POR EMPLEADO", meta)
    stats_row = meta_end + 1
    ws_sum.cell(row=stats_row, column=1, value="Días evaluados:").font = Font(bold=True)
    ws_sum.cell(row=stats_row, column=2, value=report_data.get("total_days", 0))
    ws_sum.cell(row=stats_row, column=4, value="Días con tardanza:").font = Font(bold=True)
    ws_sum.cell(row=stats_row, column=5, value=report_data.get("late_days", 0))
    ws_sum.cell(row=stats_row, column=7, value="Empleados con tardanza:").font = Font(bold=True)
    ws_sum.cell(row=stats_row, column=8, value=report_data.get("persons_with_late", 0))
    totals_row = _write_tardiness_highlight_totals(ws_sum, stats_row + 2, report_data)
    hdr_row = totals_row + 2

    person_headers = [
        "ID", "Nombre", "Sede", "Días asistió", "Días puntuales", "Días con tardanza",
        "Días tarde entrada", "Días tarde almuerzo",
        "Min. entrada", "Min. almuerzo", "Min. bruto", "Tolerancia", "Min. neto", "Promedio tardanza",
    ]
    for c, h in enumerate(person_headers, 1):
        ws_sum.cell(row=hdr_row, column=c, value=h)
    style_header_row(ws_sum, hdr_row, len(person_headers))
    _style_tardiness_minute_headers(ws_sum, hdr_row)

    persons = report_data.get("by_person", [])
    for i, p in enumerate(persons, hdr_row + 1):
        ws_sum.cell(row=i, column=1, value=p.get("user_id"))
        ws_sum.cell(row=i, column=2, value=p.get("name"))
        ws_sum.cell(row=i, column=3, value=p.get("sede"))
        ws_sum.cell(row=i, column=4, value=p.get("days_with_attendance", 0))
        ws_sum.cell(row=i, column=5, value=p.get("punctual_days", 0))
        ws_sum.cell(row=i, column=6, value=p.get("late_days", 0))
        ws_sum.cell(row=i, column=7, value=p.get("late_days_entrada", 0))
        ws_sum.cell(row=i, column=8, value=p.get("late_days_almuerzo", 0))
        ws_sum.cell(row=i, column=9, value=p.get("late_minutes_entrada", 0))
        ws_sum.cell(row=i, column=10, value=p.get("late_minutes_almuerzo", 0))
        ws_sum.cell(row=i, column=11, value=p.get("late_minutes_gross", p.get("late_minutes", 0)))
        ws_sum.cell(row=i, column=12, value=p.get("monthly_tolerance_applied", 0))
        ws_sum.cell(row=i, column=13, value=p.get("late_minutes", 0))
        ws_sum.cell(row=i, column=14, value=p.get("avg_late_minutes", 0))
        _style_tardiness_minute_row(ws_sum, i)
    if persons:
        total_row = hdr_row + len(persons) + 1
        _append_tardiness_person_totals_row(ws_sum, total_row, report_data)
    auto_width(ws_sum, min_w=10, max_w=22)

    ws_det = wb.create_sheet("DETALLE DIARIO")
    det_headers = [
        "Fecha", "ID", "Nombre", "Sede", "Hora esperada", "Entrada", "Salida almuerzo",
        "Regreso almuerzo", "Salida final", "Marcaciones", "Estado marcaciones",
        "Observación marcaciones", "Tardanza entrada", "Tardanza almuerzo", "Total min.", "Estado",
    ]
    for c, h in enumerate(det_headers, 1):
        ws_det.cell(row=1, column=c, value=h)
    style_header_row(ws_det, 1, len(det_headers))

    details = sorted(
        report_data.get("details", []),
        key=lambda x: (x.get("date", ""), x.get("user_name") or x.get("user_id", "")),
    )
    for i, d in enumerate(details, 2):
        entrada = d.get("entrada", "") or ""
        salida_alm = d.get("salida_almuerzo", "") or ""
        entrada_alm = d.get("entrada_almuerzo", "") or ""
        ws_det.cell(row=i, column=1, value=d.get("date"))
        ws_det.cell(row=i, column=2, value=d.get("user_id"))
        ws_det.cell(row=i, column=3, value=d.get("user_name"))
        ws_det.cell(row=i, column=4, value=d.get("sede_name"))
        ws_det.cell(row=i, column=5, value=d.get("hora_esperada"))
        salida = d.get("salida", "") or ""
        ws_det.cell(row=i, column=6, value=entrada.split(" ")[1] if " " in entrada else entrada)
        ws_det.cell(row=i, column=7, value=salida_alm.split(" ")[1] if " " in salida_alm else salida_alm)
        ws_det.cell(row=i, column=8, value=entrada_alm.split(" ")[1] if " " in entrada_alm else entrada_alm)
        ws_det.cell(row=i, column=9, value=salida.split(" ")[1] if " " in salida else salida)
        total_m = d.get("total_marcaciones", 0)
        esperadas = d.get("marcaciones_esperadas", 4)
        ws_det.cell(row=i, column=10, value=f"{total_m}/{esperadas}")
        ws_det.cell(row=i, column=11, value=d.get("estado_marcaciones", ""))
        ws_det.cell(row=i, column=12, value=d.get("observacion_marcaciones", ""))
        ws_det.cell(row=i, column=13, value=d.get("tardanza_minutos", 0))
        ws_det.cell(row=i, column=14, value=d.get("tardanza_almuerzo_minutos", 0))
        ws_det.cell(row=i, column=15, value=d.get("tardanza_total_minutos", 0))
        ws_det.cell(row=i, column=16, value=d.get("estado_asistencia"))
    auto_width(ws_det)

    by_uid = defaultdict(list)
    for d in details:
        by_uid[d.get("user_id")].append(d)

    for p in report_data.get("by_person", []):
        uid = p.get("user_id")
        recs = by_uid.get(uid, [])
        if not recs:
            continue
        sheet = safe_sheet_name(uid, p.get("name"), used_names)
        ws = wb.create_sheet(sheet)
        ws["A1"] = f"EMPLEADO: {p.get('name') or 'Sin nombre'}"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:L1")
        ws["A2"] = f"ID: {uid} · Sede: {p.get('sede', '—')}"
        ws["A3"] = (
            f"Días asistió: {p.get('days_with_attendance', 0)} · "
            f"Tardanza neta: {p.get('late_minutes', 0)} min "
            f"(bruto {p.get('late_minutes_gross', p.get('late_minutes', 0))}, "
            f"tolerancia {p.get('monthly_tolerance_applied', 0)} min) · "
            f"Promedio: {p.get('avg_late_minutes', 0)} min"
        )
        row = 5
        emp_headers = [
            "Fecha", "Entrada", "Esperada", "Salida almuerzo", "Regreso almuerzo",
            "Tard. entrada", "Tard. almuerzo", "Total min.", "Estado",
        ]
        for c, h in enumerate(emp_headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(emp_headers))
        for rec in recs:
            row += 1
            entrada = rec.get("entrada", "") or ""
            salida_alm = rec.get("salida_almuerzo", "") or ""
            entrada_alm = rec.get("entrada_almuerzo", "") or ""
            ws.cell(row=row, column=1, value=rec.get("date"))
            ws.cell(row=row, column=2, value=entrada.split(" ")[1] if " " in entrada else entrada)
            ws.cell(row=row, column=3, value=rec.get("hora_esperada"))
            ws.cell(row=row, column=4, value=salida_alm.split(" ")[1] if " " in salida_alm else salida_alm)
            ws.cell(row=row, column=5, value=entrada_alm.split(" ")[1] if " " in entrada_alm else entrada_alm)
            ws.cell(row=row, column=6, value=rec.get("tardanza_minutos", 0))
            ws.cell(row=row, column=7, value=rec.get("tardanza_almuerzo_minutos", 0))
            ws.cell(row=row, column=8, value=rec.get("tardanza_total_minutos", 0))
            ws.cell(row=row, column=9, value=rec.get("estado_asistencia"))
        auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def tardiness_person_csv_rows(report_data):
    headers = [
        "ID", "Nombre", "Sede", "Días asistió", "Días puntuales", "Días con tardanza",
        "Días tarde entrada", "Días tarde almuerzo",
        "Min. entrada", "Min. almuerzo", "Min. bruto", "Tolerancia mensual", "Min. neto", "Promedio tardanza",
    ]
    rows = []
    for p in report_data.get("by_person", []):
        rows.append([
            p.get("user_id"), p.get("name"), p.get("sede"),
            p.get("days_with_attendance", 0), p.get("punctual_days", 0),
            p.get("late_days", 0), p.get("late_days_entrada", 0),
            p.get("late_days_almuerzo", 0), p.get("late_minutes_entrada", 0),
            p.get("late_minutes_almuerzo", 0), p.get("late_minutes_gross", p.get("late_minutes", 0)),
            p.get("monthly_tolerance_applied", 0), p.get("late_minutes", 0),
            p.get("avg_late_minutes", 0),
        ])
    return headers, rows


def tardiness_detail_csv_rows(report_data):
    headers = [
        "Fecha", "ID", "Nombre", "Sede", "Hora esperada", "Entrada", "Salida almuerzo",
        "Regreso almuerzo", "Salida final", "Marcaciones", "Estado marcaciones",
        "Observación marcaciones", "Tardanza entrada", "Tardanza almuerzo", "Total min.", "Estado",
    ]
    rows = []
    for d in sorted(report_data.get("details", []), key=lambda x: (x.get("date", ""), x.get("user_id", ""))):
        entrada = d.get("entrada", "") or ""
        salida_alm = d.get("salida_almuerzo", "") or ""
        entrada_alm = d.get("entrada_almuerzo", "") or ""
        salida = d.get("salida", "") or ""
        total_m = d.get("total_marcaciones", 0)
        esperadas = d.get("marcaciones_esperadas", 4)
        rows.append([
            d.get("date"), d.get("user_id"), d.get("user_name"), d.get("sede_name"),
            d.get("hora_esperada"),
            entrada.split(" ")[1] if " " in entrada else entrada,
            salida_alm.split(" ")[1] if " " in salida_alm else salida_alm,
            entrada_alm.split(" ")[1] if " " in entrada_alm else entrada_alm,
            salida.split(" ")[1] if " " in salida else salida,
            f"{total_m}/{esperadas}",
            d.get("estado_marcaciones", ""),
            d.get("observacion_marcaciones", ""),
            d.get("tardanza_minutos", 0), d.get("tardanza_almuerzo_minutos", 0),
            d.get("tardanza_total_minutos", 0), d.get("estado_asistencia"),
        ])
    return headers, rows


def build_excel_monthly_tardiness(report_data, meta=None):
    meta = meta or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "TARDANZAS MENSUAL"
    meta_end = _write_meta_block(ws, "REPORTE MENSUAL DE TARDANZAS", meta)
    period_row = meta_end + 1
    ws.cell(row=period_row, column=1, value="Período").font = Font(bold=True)
    ws.cell(row=period_row, column=2, value=report_data.get("month_label") or meta.get("month_label", ""))
    ws.cell(row=period_row, column=4, value="Empleados con tardanza:").font = Font(bold=True)
    ws.cell(row=period_row, column=5, value=report_data.get("persons_with_late", 0))
    totals_row = _write_tardiness_highlight_totals(ws, period_row + 2, report_data)
    hdr_row = totals_row + 2

    headers = [
        "ID", "Nombre", "Sede", "Días laborables asistió", "Días puntuales", "Días con tardanza",
        "Días tarde entrada", "Días tarde almuerzo",
        "Min. entrada", "Min. almuerzo", "Min. bruto", "Tolerancia", "Min. neto", "Promedio tardanza",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header_row(ws, hdr_row, len(headers))
    _style_tardiness_minute_headers(ws, hdr_row)

    persons = report_data.get("by_person", [])
    for i, p in enumerate(persons, hdr_row + 1):
        ws.cell(row=i, column=1, value=p.get("user_id"))
        ws.cell(row=i, column=2, value=p.get("name"))
        ws.cell(row=i, column=3, value=p.get("sede"))
        ws.cell(row=i, column=4, value=p.get("days_with_attendance", 0))
        ws.cell(row=i, column=5, value=p.get("punctual_days", 0))
        ws.cell(row=i, column=6, value=p.get("late_days", 0))
        ws.cell(row=i, column=7, value=p.get("late_days_entrada", 0))
        ws.cell(row=i, column=8, value=p.get("late_days_almuerzo", 0))
        ws.cell(row=i, column=9, value=p.get("late_minutes_entrada", 0))
        ws.cell(row=i, column=10, value=p.get("late_minutes_almuerzo", 0))
        ws.cell(row=i, column=11, value=p.get("late_minutes_gross", p.get("late_minutes", 0)))
        ws.cell(row=i, column=12, value=p.get("monthly_tolerance_applied", 0))
        ws.cell(row=i, column=13, value=p.get("late_minutes", 0))
        ws.cell(row=i, column=14, value=p.get("avg_late_minutes", 0))
        _style_tardiness_minute_row(ws, i)
    if persons:
        total_row = hdr_row + len(persons) + 1
        _append_tardiness_person_totals_row(ws, total_row, report_data)
    auto_width(ws, min_w=10, max_w=24)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def monthly_tardiness_csv_rows(report_data):
    headers = [
        "ID", "Nombre", "Sede", "Días laborables asistió", "Días puntuales", "Días con tardanza",
        "Días tarde entrada", "Días tarde almuerzo",
        "Min. entrada", "Min. almuerzo", "Min. bruto", "Tolerancia mensual", "Min. neto", "Promedio tardanza",
    ]
    rows = []
    for p in report_data.get("by_person", []):
        rows.append([
            p.get("user_id"), p.get("name"), p.get("sede"),
            p.get("days_with_attendance", 0), p.get("punctual_days", 0),
            p.get("late_days", 0), p.get("late_days_entrada", 0),
            p.get("late_days_almuerzo", 0), p.get("late_minutes_entrada", 0),
            p.get("late_minutes_almuerzo", 0), p.get("late_minutes_gross", p.get("late_minutes", 0)),
            p.get("monthly_tolerance_applied", 0), p.get("late_minutes", 0),
            p.get("avg_late_minutes", 0),
        ])
    return headers, rows