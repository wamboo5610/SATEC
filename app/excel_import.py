"""Importación de reportes Excel exportados desde relojes ZKTeco."""

from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from . import attendance_logic as al
from . import database as db
from .paths import get_reportes_dir

REPORTES_DIR = get_reportes_dir()

HEADER_MARKERS = {"id", "nombre", "fecha / hora", "fecha/hora", "fecha y hora"}


def list_report_files() -> list[dict]:
    if not REPORTES_DIR.exists():
        REPORTES_DIR.mkdir(parents=True, exist_ok=True)
        return []
    files = []
    for path in sorted(REPORTES_DIR.glob("*.xlsx")):
        stat = path.stat()
        files.append({
            "filename": path.name,
            "size_kb": round(stat.st_size / 1024, 1),
            "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
        })
    return files


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | tuple[None, None]:
    for idx, row in enumerate(rows):
        mapping: dict[str, int] = {}
        for col, cell in enumerate(row):
            key = _normalize_header(cell)
            if key in HEADER_MARKERS or key == "estado":
                mapping[key] = col
        if "id" in mapping and any(k in mapping for k in ("fecha / hora", "fecha/hora", "fecha y hora")):
            return idx, mapping
    return None, None


def _col(mapping: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in mapping:
            return mapping[name]
    return None


def parse_datetime(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = str(value).strip()
    if not text:
        return None
    for fmt in (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
    ):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return None


def parse_zk_excel(
    file_bytes: bytes,
    sede_name: str | None = None,
    source_label: str = "excel",
) -> tuple[list[dict], dict]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx, mapping = _find_header_row(rows)
    if header_idx is None or mapping is None:
        raise ValueError("No se encontró la fila de encabezados (ID, Nombre, Fecha / Hora)")

    col_id = _col(mapping, "id")
    col_name = _col(mapping, "nombre")
    col_ts = _col(mapping, "fecha / hora", "fecha/hora", "fecha y hora")
    col_status = _col(mapping, "estado")

    if col_id is None or col_ts is None:
        raise ValueError("El Excel no contiene las columnas ID y Fecha / Hora")

    records: list[dict] = []
    skipped = 0
    users: set[str] = set()
    dates: list[str] = []

    for row in rows[header_idx + 1:]:
        if not row or col_id >= len(row):
            continue
        user_id = str(row[col_id]).strip() if row[col_id] is not None else ""
        if not user_id or user_id.lower() == "none":
            skipped += 1
            continue
        timestamp = parse_datetime(row[col_ts] if col_ts < len(row) else None)
        if not timestamp:
            skipped += 1
            continue
        user_name = ""
        if col_name is not None and col_name < len(row) and row[col_name]:
            user_name = str(row[col_name]).strip()
        estado = ""
        if col_status is not None and col_status < len(row) and row[col_status]:
            estado = str(row[col_status]).strip().lower()
        status = 0
        if "salida" in estado:
            status = 1
        records.append({
            "user_id": user_id,
            "user_name": user_name,
            "timestamp": timestamp,
            "status": status,
            "verify_mode": 0,
            "device_serial": None,
            "source": source_label,
            "sede_name": sede_name or "Excel importado",
            "sede_id": None,
        })
        users.add(user_id)
        dates.append(timestamp[:10])

    if not records:
        raise ValueError("No se encontraron marcaciones válidas en el archivo")

    dates.sort()
    meta = {
        "total_records": len(records),
        "total_persons": len(users),
        "date_from": dates[0],
        "date_to": dates[-1],
        "skipped_rows": skipped,
        "sede_name": sede_name or "Excel importado",
        "source": source_label,
    }
    return records, meta


def calculate_tardiness_from_records(
    records: list[dict],
    schedule: dict | None = None,
    sede_name: str | None = None,
    sede_id: int | None = None,
) -> dict:
    sched = al.normalize_schedule(schedule)
    sede_label = sede_name or (records[0].get("sede_name") if records else "Excel importado")
    sid = sede_id if sede_id is not None else 0
    for r in records:
        r["sede_name"] = sede_label
        r["sede_id"] = sid
    enriched = al.infer_punch_types(records)
    emp_schedules = db.get_schedules_by_user_id()
    global_dates, by_sede = db.get_holiday_dates_lookup()
    report = al.build_tardiness_report(
        enriched, {sid: sched}, {sede_label: sid}, emp_schedules, global_dates, by_sede
    )
    return al.aggregate_tardiness_report(report)


def calculate_tardiness_from_excel(
    file_bytes: bytes,
    sede_name: str | None = None,
    schedule: dict | None = None,
    source_label: str = "excel",
    sede_id: int | None = None,
) -> dict:
    records, meta = parse_zk_excel(file_bytes, sede_name=sede_name, source_label=source_label)
    report = calculate_tardiness_from_records(
        records, schedule=schedule, sede_name=sede_name, sede_id=sede_id
    )
    report["meta"] = meta
    return report


def read_local_report(filename: str) -> bytes:
    safe_name = Path(filename).name
    path = REPORTES_DIR / safe_name
    if not path.exists() or path.suffix.lower() != ".xlsx":
        raise FileNotFoundError(f"Archivo no encontrado: {safe_name}")
    return path.read_bytes()


def excel_device_serial(sede_id: int | None, filename: str = "") -> str:
    if sede_id:
        return f"EXCEL-S{sede_id}"
    stem = re.sub(r"[^A-Za-z0-9]", "", Path(filename).stem)[:20] if filename else ""
    return f"EXCEL-{stem.upper() or 'IMPORT'}"


def extract_users_from_records(records: list[dict]) -> list[dict]:
    users: dict[str, dict] = {}
    for r in records:
        uid = str(r.get("user_id", ""))
        if not uid:
            continue
        if r.get("user_name"):
            users[uid] = {"user_id": uid, "name": r["user_name"], "privilege": ""}
        elif uid not in users:
            users[uid] = {"user_id": uid, "name": "", "privilege": ""}
    return list(users.values())


def prepare_records_for_db(
    records: list[dict],
    device_serial: str,
    sede_id: int | None = None,
) -> list[dict]:
    enriched = al.infer_punch_types([dict(r) for r in records])
    for r in enriched:
        r["device_serial"] = device_serial
        r["source"] = "excel"
        if sede_id is not None:
            r["sede_id"] = sede_id
    return enriched


def import_excel_to_db(
    file_bytes: bytes,
    sede_id: int,
    sede_name: str | None = None,
    filename: str = "",
    *,
    replace_existing: bool = False,
) -> dict:
    sedes = {s["id"]: s["name"] for s in db.get_sedes()}
    if sede_id not in sedes:
        raise ValueError("Sede no encontrada")
    label = (sede_name or sedes[sede_id]).strip()
    device_serial = excel_device_serial(sede_id, filename)
    removed_previous = 0
    if replace_existing:
        removed_previous = db.delete_attendance_by_excel_sede(sede_id)
    records, meta = parse_zk_excel(file_bytes, sede_name=label, source_label="excel")
    device_id = db.ensure_excel_device(sede_id, label, device_serial)
    enriched = prepare_records_for_db(records, device_serial, sede_id)
    users = extract_users_from_records(enriched)
    db.upsert_users(users, device_serial)
    inserted = db.insert_attendance(enriched)
    return {
        "ok": True,
        "records_total": len(enriched),
        "records_new": inserted,
        "records_duplicate": len(enriched) - inserted,
        "records_removed_previous": removed_previous,
        "users_updated": len(users),
        "device_serial": device_serial,
        "device_id": device_id,
        "sede_id": sede_id,
        "sede_name": label,
        "meta": meta,
    }